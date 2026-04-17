#!/usr/bin/env python3
"""
AutoChecker — LLM-powered math grading CLI (Codex backend).

Runs grading through OpenAI's Codex CLI (`codex exec`) so calls go
against your ChatGPT subscription quota instead of per-token API billing.

Requires:
  - `codex` CLI installed (https://developers.openai.com/codex/cli)
  - One-time `codex login` with your ChatGPT account

Run `autochecker` in any directory containing:
  - A solutions file (*solutions*.pdf)
  - Student submission files (*.jpg, *.png, *.pdf)
"""

import csv
import json
import shutil
import subprocess
import sys
import tempfile
import threading
import time
from collections import defaultdict
from datetime import datetime, timezone
from importlib.metadata import PackageNotFoundError, version as _pkg_version
from pathlib import Path
from typing import Callable

import fitz  # pymupdf
import openpyxl
from rich.console import Console, Group
from rich.live import Live
from rich.panel import Panel
from rich.progress import (
    Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn,
    TimeElapsedColumn, ProgressColumn,
)
from rich.prompt import Prompt, Confirm
from rich.table import Table
from rich.text import Text
from rich.theme import Theme


def _version() -> str:
    try:
        return _pkg_version("autochecker")
    except PackageNotFoundError:
        return "dev"


VERSION = _version()

# ── Config ───────────────────────────────────────────────────────────

CONFIG_DIR = Path.home() / ".config" / "autochecker"
CONFIG_FILE = CONFIG_DIR / "config.json"

DEFAULT_CONFIG = {
    # None → let Codex pick the default model for your ChatGPT plan.
    # Override with e.g. "gpt-5-codex" if you want a specific one.
    "default_model": None,
    "codex_timeout_seconds": 1800,
}


def load_config() -> dict:
    config = dict(DEFAULT_CONFIG)
    if CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE) as f:
                config.update(json.load(f))
        except (json.JSONDecodeError, OSError):
            pass
    return config


def save_config(config: dict):
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    with open(CONFIG_FILE, "w") as f:
        json.dump(config, f, indent=2)


# ── Theme & console ──────────────────────────────────────────────────

theme = Theme({
    "info": "cyan",
    "success": "green",
    "warning": "yellow",
    "error": "red bold",
    "header": "bold magenta",
    "muted": "dim",
    "score.high": "bold green",
    "score.mid": "bold yellow",
    "score.low": "bold red",
})
console = Console(theme=theme)

# ── Rubric / grading prompts (built dynamically from the solutions file) ──

# Fallback rubric if auto-detection from the solutions file fails.
DEFAULT_RUBRIC: dict[str, int] = {"1": 2, "2": 2, "3": 3, "4": 2, "5": 3}


def make_grading_schema(rubric: dict[str, int]) -> dict:
    """Build a JSON Schema for a grading response from an ordered rubric."""
    return {
        "type": "object",
        "properties": {
            "scores": {
                "type": "object",
                "properties": {
                    qid: {"type": "integer", "minimum": 0, "maximum": mx}
                    for qid, mx in rubric.items()
                },
                "required": list(rubric.keys()),
                "additionalProperties": False,
            },
            "total": {"type": "integer", "minimum": 0,
                      "maximum": sum(rubric.values())},
            "notes": {"type": "string"},
        },
        "required": ["scores", "total", "notes"],
        "additionalProperties": False,
    }


def make_grading_prompt(rubric: dict[str, int]) -> str:
    total = sum(rubric.values())
    return f"""\
You are grading a student's handwritten submission against the provided \
official solutions.

The problem set has {len(rubric)} question(s). \
Point values per question: {json.dumps(rubric)}. Total: {total} points.

The attached images contain:
  1. The official worked solutions (first pages).
  2. The student's handwritten submission (remaining pages).

GRADING INSTRUCTIONS:
- For each question the student attempted, compare their work against the official solution.
- Award points based on correctness and completeness. Partial credit is allowed.
- If a question is not attempted, give 0 points.
- Be fair but rigorous. Minor notation differences are OK. The key steps must be present.

Return ONLY the JSON object matching the provided schema. Do not run tools, do not edit files, \
do not ask clarifying questions.
"""


RUBRIC_SCHEMA = {
    "type": "object",
    "properties": {
        "questions": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "id": {"type": "string"},
                    "points": {"type": "integer", "minimum": 1, "maximum": 100},
                },
                "required": ["id", "points"],
                "additionalProperties": False,
            },
            "minItems": 1,
            "maxItems": 30,
        },
    },
    "required": ["questions"],
    "additionalProperties": False,
}

RUBRIC_PROMPT = """\
Look at the attached worked-solutions pages and extract the **grading rubric** \
— i.e. the criteria a teacher would use to mark the work. This is usually \
listed in a table, scoring guide, or criteria section at the end of the \
solutions. It is NOT the same as the list of derivation steps inside the \
solution itself.

What to look for (in order of preference):
  1. An explicit scoring table / criteria list (e.g. "Критерии оценивания", \
"Grading rubric", a table of per-question items with points).
  2. Labelled sub-parts of each question (part a, part b, 1.1, 1.2, etc.) with \
points next to each.
  3. Only if none of the above exists: use each top-level question as a single \
item with its maximum points.

For each rubric item return:
  - "id":     the label as it appears in the document. Use dotted notation \
for sub-parts: "1.1", "1.2", "2a". Plain "1", "2" for items with no sub-parts.
  - "points": the maximum points for that rubric item (positive integer)

Return JSON of the form:
  {"questions": [
      {"id": "1.1", "points": 4},
      {"id": "1.2", "points": 6},
      {"id": "2",   "points": 8},
      ...
  ]}

Do NOT list every derivation step in the worked solution as a separate rubric \
item. Only list items the grading guide actually scores. Preserve the order \
items appear. Do not run tools, do not ask questions.
"""

NAME_DETECT_SCHEMA = {
    "type": "object",
    "properties": {
        "name": {"type": ["string", "null"]},
    },
    "required": ["name"],
    "additionalProperties": False,
}

NAME_DETECT_PROMPT = """\
Look at the attached scanned student submission pages. Your only job is to \
find the student's name or signature if they wrote one anywhere on the pages \
(header, corner, signature line, etc.).

Return JSON matching the schema:
  {"name": "<full name as written>"}   if a name is clearly visible
  {"name": null}                       if no name or it's unreadable

Keep the original script (Cyrillic stays Cyrillic). Do not guess from the \
filename. Do not run tools.
"""

NAME_MATCH_SCHEMA = {
    "type": "object",
    "properties": {
        "matches": {
            "type": "object",
            "additionalProperties": {
                "oneOf": [
                    {"type": "array", "items": {"type": "string"}, "minItems": 2, "maxItems": 2},
                    {"type": "null"},
                ]
            },
        }
    },
    "required": ["matches"],
    "additionalProperties": False,
}

NAME_MATCHING_PROMPT = """\
Match student submission filename prefixes to the spreadsheet roster.

Submission prefixes:
{file_groups}

Spreadsheet roster (Name, Surname):
{roster}

Filename prefixes are abbreviated FirstName+LastName (e.g., "AlexKto" → "Alexandros Ktoris"). \
For each prefix, find the best matching student from the roster, or null if no good match exists.

Return ONLY a JSON object of the form {{"matches": {{"Prefix": ["FirstName", "LastName"] or null, ...}}}}. \
Do not run tools, do not edit files.
"""


# ── PDF rasterization ────────────────────────────────────────────────

def render_pdf_pages(pdf_path: Path, out_dir: Path, prefix: str) -> list[Path]:
    """Rasterize every page of a PDF to PNG files in out_dir."""
    doc = fitz.open(pdf_path)
    paths = []
    for i, page in enumerate(doc, 1):
        pix = page.get_pixmap(dpi=200)
        out = out_dir / f"{prefix}_p{i:02d}.png"
        pix.save(str(out))
        paths.append(out)
    doc.close()
    return paths


def materialize_submission(files: list[Path], out_dir: Path, prefix: str) -> list[Path]:
    """Return on-disk image paths for a student's submission, rasterizing PDFs as needed."""
    result = []
    for i, f in enumerate(files, 1):
        suf = f.suffix.lower()
        if suf == ".pdf":
            result.extend(render_pdf_pages(f, out_dir, f"{prefix}_sub{i}"))
        elif suf in (".jpg", ".jpeg", ".png"):
            result.append(f)
        else:
            raise ValueError(f"Unsupported file: {f}")
    return result


# ── Directory scanning ───────────────────────────────────────────────

def find_solutions_file(directory: Path) -> Path | None:
    for f in directory.iterdir():
        if f.suffix.lower() == ".pdf" and "solution" in f.name.lower():
            return f
    return None


def _collect_submission_files(directory: Path, solutions_file: Path) -> list[Path]:
    out = []
    for f in sorted(directory.iterdir()):
        if f == solutions_file:
            continue
        if f.suffix.lower() not in (".jpg", ".jpeg", ".png", ".pdf"):
            continue
        if f.name.startswith("."):
            continue
        out.append(f)
    return out


def _autodetect_grouping(files: list[Path]) -> str:
    """Pick 'prefix' (many students × multi-page) vs 'per-file' (one-PDF-per-student).

    Heuristic: if all files collapse to a single stripped-digits prefix AND
    there are at least 5 files, each file is almost certainly a separate
    student (e.g., `img-<timestamp>.pdf`). Otherwise use prefix grouping.
    """
    if len(files) < 5:
        return "prefix"
    prefixes = {(f.stem.rstrip("0123456789") or f.stem) for f in files}
    return "per-file" if len(prefixes) == 1 else "prefix"


def find_submissions(directory: Path, solutions_file: Path,
                     mode: str = "auto") -> tuple[dict[str, list[Path]], str]:
    """Return (groups, resolved_mode).

    `mode` is one of 'auto', 'prefix', 'per-file'. With 'auto', the mode is
    chosen by `_autodetect_grouping`. The resolved mode is returned so the
    caller can display it.
    """
    files = _collect_submission_files(directory, solutions_file)
    resolved = _autodetect_grouping(files) if mode == "auto" else mode

    groups: dict[str, list[Path]] = defaultdict(list)
    if resolved == "per-file":
        for f in files:
            groups[f.stem].append(f)
    else:  # 'prefix'
        for f in files:
            stem = f.stem
            prefix = stem.rstrip("0123456789") or stem
            groups[prefix].append(f)
    return dict(groups), resolved


def find_spreadsheet(directory: Path) -> Path | None:
    """Return the first .xlsx file in `directory`, ignoring Excel temp
    files (`~$*.xlsx`). Only the current directory is searched — parents
    are not, to avoid silently picking up unrelated rosters."""
    try:
        candidates = sorted(
            p for p in directory.glob("*.xlsx")
            if not p.name.startswith("~$") and not p.name.startswith(".")
        )
    except OSError:
        candidates = []
    return candidates[0] if candidates else None


# ── Codex wrapper ────────────────────────────────────────────────────

def check_codex_available():
    if shutil.which("codex") is None:
        console.print("[error]`codex` CLI not found.[/] Install from https://developers.openai.com/codex/cli")
        console.print("[muted]Then run `codex login` once to authenticate with your ChatGPT account.[/]")
        sys.exit(1)


def codex_exec(prompt: str, image_paths: list[Path], schema: dict, model: str | None,
               workdir: Path, timeout: int, tag: str,
               on_event: Callable[[dict], None] | None = None) -> dict:
    """Run `codex exec` non-interactively and return the parsed JSON result.

    With `--json`, Codex prints JSONL events as it works. Each event is passed
    to `on_event` so the caller can surface live progress (reasoning steps,
    token counts, etc.). The final JSON response is still read from the file
    written by `--output-last-message` once the process exits.
    """
    schema_path = workdir / f"schema_{tag}.json"
    schema_path.write_text(json.dumps(schema))
    result_path = workdir / f"result_{tag}.json"
    if result_path.exists():
        result_path.unlink()

    # Prompt must precede `-i` flags: clap's `-i <FILE>...` otherwise swallows
    # the positional prompt as another image path.
    cmd = [
        "codex", "exec",
        "--json",
        "--sandbox", "read-only",
        "--skip-git-repo-check",
        "--ephemeral",
        "--color", "never",
        # Surface the model's internal reasoning as JSONL events so the UI
        # can display live "thoughts" while a call is in flight.
        "-c", "show_raw_agent_reasoning=true",
        "-c", 'model_reasoning_summary="auto"',
        "--output-schema", str(schema_path),
        "--output-last-message", str(result_path),
    ]
    if model:
        cmd.extend(["-m", model])
    cmd.append(prompt)
    if image_paths:
        # Resolve to absolute: cwd is the ephemeral tmpdir, not the caller's cwd.
        abs_imgs = [str(Path(p).resolve()) for p in image_paths]
        cmd.extend(["-i", ",".join(abs_imgs)])

    proc = subprocess.Popen(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,   # merge so we read a single stream
        text=True,
        bufsize=1,                  # line-buffered
        cwd=str(workdir),
        stdin=subprocess.DEVNULL,
    )

    # Hard timeout — kill the process if it takes too long.
    timed_out = {"flag": False}
    def _kill():
        timed_out["flag"] = True
        try:
            proc.kill()
        except Exception:
            pass
    timer = threading.Timer(timeout, _kill)
    timer.daemon = True
    timer.start()

    non_json: list[str] = []
    try:
        assert proc.stdout is not None
        for raw_line in proc.stdout:
            line = raw_line.rstrip("\n")
            if not line:
                continue
            try:
                event = json.loads(line)
            except json.JSONDecodeError:
                non_json.append(line)
                continue
            if on_event is not None:
                try:
                    on_event(event)
                except Exception:
                    pass  # never let a UI callback kill the subprocess loop
        rc = proc.wait()
    finally:
        timer.cancel()

    if timed_out["flag"]:
        raise RuntimeError(f"codex exec timed out after {timeout}s")
    if rc != 0:
        tail = " | ".join(non_json[-5:])[:400] or f"exit code {rc}"
        raise RuntimeError(f"codex exec failed (rc={rc}): {tail}")

    if not result_path.exists():
        raise RuntimeError("codex exec produced no result file")

    raw = result_path.read_text().strip()
    if raw.startswith("```"):
        raw = raw.split("\n", 1)[1] if "\n" in raw else raw[3:]
        if raw.endswith("```"):
            raw = raw[:-3]
        raw = raw.strip()

    try:
        return json.loads(raw)
    except json.JSONDecodeError as e:
        raise RuntimeError(f"codex exec returned invalid JSON: {e}; raw={raw[:200]}")


def event_phase(event: dict, counters: dict) -> str | None:
    """Map a Codex JSONL event to a short human phase label, or None to skip.

    `counters` is a per-call dict used to accumulate things like
    reasoning-step count across events.
    """
    et = event.get("type", "")
    if et == "thread.started":
        return "session opened"
    if et == "turn.started":
        return "thinking"
    if et in ("item.started", "item.completed"):
        item = event.get("item") or {}
        itype = item.get("type", "")
        if itype == "reasoning":
            if et == "item.completed":
                counters["reasoning"] = counters.get("reasoning", 0) + 1
                return f"reasoning · step {counters['reasoning']}"
            return "reasoning"
        if itype == "agent_message":
            return "writing answer"
        if itype in ("function_call", "exec_command"):
            return "tool call"
        return None
    if et == "turn.completed":
        usage = event.get("usage") or {}
        out_tok = usage.get("output_tokens")
        return f"finishing · {out_tok} tok" if out_tok is not None else "finishing"
    return None


def event_thought(event: dict, max_len: int = 220) -> str | None:
    """Pull a single-line excerpt of reasoning text out of a reasoning event.

    Codex emits reasoning as `item.completed` events with `item.type ==
    "reasoning"` and a markdown-ish `text` field. We pick the last
    meaningful line, strip markdown decoration, and trim to `max_len`.
    """
    if event.get("type") != "item.completed":
        return None
    item = event.get("item") or {}
    if item.get("type") != "reasoning":
        return None
    text = (item.get("text") or "").strip()
    if not text:
        return None
    # Take the last non-trivial line (skip pure markdown headers).
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    pick = None
    for line in reversed(lines):
        cleaned = line.strip("*").strip("#").strip()
        if cleaned:
            pick = cleaned
            break
    if pick is None:
        pick = lines[-1]
    # Collapse whitespace, strip stray markdown tokens.
    pick = " ".join(pick.split())
    pick = pick.replace("**", "").replace("`", "")
    if len(pick) > max_len:
        pick = pick[: max_len - 1].rstrip() + "…"
    return pick


# ── Grading ──────────────────────────────────────────────────────────

def grade_student(model: str | None, timeout: int, solutions_paths: list[Path],
                  student_prefix: str, submission_paths: list[Path],
                  workdir: Path, rubric: dict[str, int],
                  on_event: Callable[[dict], None] | None = None) -> dict:
    prompt = make_grading_prompt(rubric) + f"\n\nStudent identifier: {student_prefix}"
    schema = make_grading_schema(rubric)
    images = solutions_paths + submission_paths
    empty_scores = {qid: 0 for qid in rubric}
    try:
        result = codex_exec(prompt, images, schema, model, workdir, timeout,
                            tag=f"grade_{student_prefix}", on_event=on_event)
    except RuntimeError as e:
        return {
            "scores": dict(empty_scores),
            "total": 0,
            "notes": f"ERROR: {str(e)[:200]}",
        }
    scores = result.get("scores") or {}
    # Coerce and fill missing keys with 0 so the table never has gaps.
    result["scores"] = {qid: int(scores.get(qid, 0)) for qid in rubric}
    result.setdefault("total", sum(result["scores"].values()))
    result.setdefault("notes", "")
    return result


import re as _re


def parse_rubric_pattern(spec: str) -> dict[str, int] | None:
    """Parse shorthand rubric patterns typed by the user:

        5x4@2      → 5 questions × 4 sub-parts × 2 pts each   (20 items, 40 pts)
        5x4        → 5 × 4 with 1 pt each                     (20 items, 20 pts)
        5@16       → 5 top-level questions × 16 pts each      (5 items,  80 pts)
        5          → 5 top-level × 1 pt each                  (5 items,   5 pts)
    """
    spec = spec.strip().lower()
    if not spec:
        return None
    m = _re.fullmatch(r"(\d+)x(\d+)(?:@(\d+))?", spec)
    if m:
        n, k = int(m.group(1)), int(m.group(2))
        p = int(m.group(3)) if m.group(3) else 1
        if n <= 0 or k <= 0 or p <= 0:
            return None
        rubric: dict[str, int] = {}
        for q in range(1, n + 1):
            for sub in range(1, k + 1):
                rubric[f"{q}.{sub}"] = p
        return rubric
    m = _re.fullmatch(r"(\d+)(?:@(\d+))?", spec)
    if m:
        n = int(m.group(1))
        p = int(m.group(2)) if m.group(2) else 1
        if n <= 0 or p <= 0:
            return None
        return {str(q): p for q in range(1, n + 1)}
    return None


def prompt_rubric_override() -> dict[str, int] | None:
    """Ask the user for a manual rubric pattern. Returns None on cancel."""
    console.print()
    console.print("  [muted]Enter a rubric pattern, or press Enter to abort:[/]")
    console.print("    [bold]NxM@P[/]  — N questions × M parts × P pts each "
                  "[muted](e.g. 5x4@2)[/]")
    console.print("    [bold]NxM[/]    — N × M parts × 1 pt each")
    console.print("    [bold]N@P[/]    — N top-level questions × P pts each")
    console.print("    [bold]N[/]      — N top-level questions × 1 pt each")
    while True:
        spec = Prompt.ask("  [bold]Rubric[/]", default="").strip()
        if not spec:
            return None
        rubric = parse_rubric_pattern(spec)
        if rubric:
            return rubric
        console.print(f"  [warning]Unrecognised pattern:[/] {spec}")


def detect_rubric(model: str | None, timeout: int,
                  solutions_paths: list[Path], workdir: Path,
                  on_event: Callable[[dict], None] | None = None
                  ) -> dict[str, int] | None:
    """Ask Codex to read the solutions PDF and extract a question → points map.

    Returns an ordered dict (insertion preserved), or None if extraction
    fails or looks suspicious (empty / non-positive values).
    """
    try:
        result = codex_exec(RUBRIC_PROMPT, solutions_paths, RUBRIC_SCHEMA,
                            model, workdir, timeout, tag="rubric",
                            on_event=on_event)
    except RuntimeError:
        return None
    questions = result.get("questions") or []
    rubric: dict[str, int] = {}
    for q in questions:
        qid = q.get("id")
        pts = q.get("points")
        if qid and isinstance(pts, int) and pts > 0:
            rubric[str(qid).strip()] = int(pts)
    return rubric or None


def detect_student_name(model: str | None, timeout: int,
                        submission_paths: list[Path],
                        workdir: Path, tag: str,
                        on_event: Callable[[dict], None] | None = None) -> str | None:
    """Quick pre-check: ask Codex to read a signature off the submission.
    Returns the detected name string, or None if none is readable/present."""
    try:
        result = codex_exec(
            NAME_DETECT_PROMPT, submission_paths, NAME_DETECT_SCHEMA,
            model, workdir, timeout, tag=f"detect_{tag}", on_event=on_event,
        )
    except RuntimeError:
        return None
    name = result.get("name")
    if not name or not isinstance(name, str):
        return None
    name = name.strip()
    return name or None


def match_names(model: str, timeout: int, file_prefixes: list[str],
                roster: list[tuple[str, str]], workdir: Path,
                signed_names: dict[str, str | None] | None = None,
                ) -> dict[str, tuple[str, str] | None]:
    roster_str = "\n".join(f"  - {n} {s}" for n, s in roster)
    if signed_names:
        groups_str = "\n".join(
            f"  - {p}" + (f"  (signed: {signed_names[p]})" if signed_names.get(p) else "")
            for p in file_prefixes
        )
    else:
        groups_str = "\n".join(f"  - {p}" for p in file_prefixes)
    prompt = NAME_MATCHING_PROMPT.format(file_groups=groups_str, roster=roster_str)
    try:
        result = codex_exec(prompt, [], NAME_MATCH_SCHEMA, model, workdir, timeout,
                            tag="name_match")
    except RuntimeError:
        return {p: None for p in file_prefixes}

    mapping = result.get("matches", {}) or {}
    out = {}
    for p in file_prefixes:
        m = mapping.get(p)
        if m and isinstance(m, list) and len(m) == 2:
            out[p] = (str(m[0]), str(m[1]))
        else:
            out[p] = None
    return out


# ── Spreadsheet ──────────────────────────────────────────────────────

def _col_letter(idx: int) -> str:
    """0-based column index → Excel-style letter (0 → A, 1 → B, ...)."""
    n = idx
    letters = ""
    while True:
        letters = chr(ord("A") + n % 26) + letters
        n = n // 26 - 1
        if n < 0:
            break
    return letters


NAME_HEADERS = ("name", "first name", "firstname", "given", "имя")
SURNAME_HEADERS = ("surname", "last name", "lastname", "family", "фамилия")
SCORE_HEADERS = ("score", "total", "grade", "points", "mark",
                 "оценка", "балл", "итог")

# Fallback column indices (0-based) used when headers can't be detected.
FALLBACK_NAME_COL = 1      # column B
FALLBACK_SURNAME_COL = 2   # column C
FALLBACK_SCORE_COL = 5     # column F


def _detect_roster_columns(ws) -> tuple[int, int, int]:
    """Inspect row 1 for name/surname/score headers. Return 0-based indices;
    fall back to B/C/F if a header can't be matched."""
    name_col = surname_col = score_col = None
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        text = str(cell).strip().lower()
        if name_col is None and any(h == text or h in text for h in NAME_HEADERS):
            # Prefer exact match for "name" over substrings like "surname".
            if "surname" in text or "last" in text or "family" in text or "фамилия" in text:
                pass
            else:
                name_col = i
        if surname_col is None and any(h in text for h in SURNAME_HEADERS):
            surname_col = i
        if score_col is None and any(h in text for h in SCORE_HEADERS):
            score_col = i
    return (
        FALLBACK_NAME_COL if name_col is None else name_col,
        FALLBACK_SURNAME_COL if surname_col is None else surname_col,
        FALLBACK_SCORE_COL if score_col is None else score_col,
    )


def _pick_roster_sheet(wb, preferred: str | None = None):
    """Return the worksheet to use as roster. Prefer `preferred` if given,
    otherwise the first sheet whose row-1 has a detectable name header,
    otherwise the first sheet."""
    if preferred and preferred in wb.sheetnames:
        return wb[preferred]
    for name in wb.sheetnames:
        ws = wb[name]
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        for cell in header:
            if cell is None:
                continue
            text = str(cell).strip().lower()
            if any(h in text for h in NAME_HEADERS + SURNAME_HEADERS):
                return ws
    return wb[wb.sheetnames[0]]


def read_roster(spreadsheet_path: Path, sheet_name: str | None = None
                ) -> tuple[list[tuple[str, str]], str, tuple[int, int, int]]:
    """Return (roster rows, resolved sheet name, (name/surname/score cols))."""
    wb = openpyxl.load_workbook(spreadsheet_path)
    try:
        ws = _pick_roster_sheet(wb, sheet_name)
        name_col, surname_col, score_col = _detect_roster_columns(ws)
        roster = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if max(name_col, surname_col) >= len(row):
                continue
            name = row[name_col].value
            surname = row[surname_col].value
            if name and surname:
                roster.append((str(name), str(surname)))
        return roster, ws.title, (name_col, surname_col, score_col)
    finally:
        wb.close()


def write_scores(spreadsheet_path: Path, scores_by_name: dict[tuple[str, str], int],
                 sheet_name: str, cols: tuple[int, int, int]) -> int:
    name_col, surname_col, score_col = cols
    wb = openpyxl.load_workbook(spreadsheet_path)
    try:
        ws = wb[sheet_name]
        filled = 0
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if max(name_col, surname_col) >= len(row):
                continue
            name = row[name_col].value
            surname = row[surname_col].value
            if name and surname:
                key = (str(name), str(surname))
                if key in scores_by_name:
                    # Extend the row if the score column is beyond current width.
                    if score_col >= len(row):
                        ws.cell(row=row[0].row, column=score_col + 1,
                                value=scores_by_name[key])
                    else:
                        row[score_col].value = scores_by_name[key]
                    filled += 1
        wb.save(spreadsheet_path)
        return filled
    finally:
        wb.close()


# ── UI ───────────────────────────────────────────────────────────────

class AnimatedPhaseColumn(ProgressColumn):
    """TextColumn replacement that breathes ellipses after active phase words.

    Cycles through a slow 4-frame ellipsis ("", "·", "··", "···") at ~0.5 Hz
    after recognisable in-flight phrases ("thinking", "reasoning", "writing
    answer", "finishing"). Idle descriptions render unchanged.
    """

    ACTIVE_PHRASES = ("writing answer", "reasoning", "finishing", "thinking")
    # Every frame is the same visible width so the progress bar to the right
    # doesn't shift as the ellipsis breathes.
    FRAMES = ("      ", " ·    ", " · ·  ", " · · ·")
    FRAMES_PER_SECOND = 2  # 0.5 s per frame → 2 s full cycle

    def render(self, task):
        desc = str(task.description or "")
        frame = int(time.monotonic() * self.FRAMES_PER_SECOND) % len(self.FRAMES)
        dots = self.FRAMES[frame]
        for phrase in self.ACTIVE_PHRASES:
            if phrase in desc:
                desc = desc.replace(phrase, f"{phrase}{dots}", 1)
                break
        return Text.from_markup(desc)


def score_style(score: int, total: int) -> str:
    ratio = score / total if total else 0
    if ratio >= 0.8:
        return "score.high"
    if ratio >= 0.5:
        return "score.mid"
    return "score.low"


KNOWN_MODELS = [
    (None, "plan default (recommended — Codex picks per your ChatGPT plan)"),
    ("gpt-5-codex", "gpt-5-codex — coding-tuned variant"),
    ("gpt-5.4", "gpt-5.4 — current Codex default on most plans"),
]


def _model_label(model: str | None) -> str:
    return model if model else "plan default"


def render_header(state: dict):
    lines = Text()
    lines.append("  autochecker ", style="bold cyan")
    lines.append(f"v{VERSION}", style="muted")
    lines.append("   math grading · Codex CLI\n\n", style="muted")

    lines.append("  dir      ", style="muted")
    lines.append(f"{state['cwd']}\n", style="info")

    lines.append("  model    ", style="muted")
    lines.append(_model_label(state["model"]), style="bold")
    lines.append("  (via ChatGPT subscription)\n", style="muted")

    sol = state["solutions_file"]
    lines.append("  crit     ", style="muted")
    if sol:
        lines.append(f"{sol.name}\n", style="success")
    else:
        lines.append("none found\n", style="warning")

    groups = state["groups"]
    lines.append("  students ", style="muted")
    if groups:
        n_files = sum(len(f) for f in groups.values())
        mode_hint = "one-per-file" if state.get("grouping") == "per-file" else "by name prefix"
        lines.append(
            f"{len(groups)} students · {n_files} files  ",
            style="success",
        )
        lines.append(f"({mode_hint})\n", style="muted")
    else:
        lines.append("none found\n", style="warning")

    sh = state["spreadsheet"]
    lines.append("  roster   ", style="muted")
    if sh:
        lines.append(f"{sh.name}", style="success")
    else:
        lines.append("not found (name matching disabled)", style="warning")

    console.print(Panel(lines, border_style="cyan", padding=(0, 0)))
    console.print(
        "  [muted]Commands:[/] "
        "[bold]/students[/] · [bold]/crit[/] · [bold]/model[/] · "
        "[bold]/group[/] · [bold]/grade[/] · [bold]/rescan[/] · "
        "[bold]/help[/] · [bold]/quit[/]"
    )


def cmd_help(state: dict, *args: str):
    tbl = Table(show_header=False, box=None, padding=(0, 2), pad_edge=False)
    tbl.add_column(style="bold cyan")
    tbl.add_column()
    tbl.add_row("/students", "list discovered student submissions")
    tbl.add_row("/crit", "show path to the solutions/criteria file")
    tbl.add_row("/model", "pick the Codex model (or save a default)")
    tbl.add_row("/group", "switch grouping (prefix / per-file / auto)")
    tbl.add_row("/grade", "run grading against the solutions file")
    tbl.add_row("/results", "re-show saved results (add 'full' for sub-parts)")
    tbl.add_row("/rescan", "re-scan the current directory")
    tbl.add_row("/status", "re-print the header")
    tbl.add_row("/help", "show this help")
    tbl.add_row("/quit", "exit (also Ctrl-D)")
    console.print()
    console.print(tbl)


def cmd_students(state: dict, *args: str):
    groups = state["groups"]
    if not groups:
        console.print("  [warning]No student submissions found in[/] " + str(state["cwd"]))
        return
    tbl = Table(border_style="cyan", title="Students", title_style="bold cyan")
    tbl.add_column("#", style="muted", width=3, justify="right")
    tbl.add_column("Prefix", style="bold")
    tbl.add_column("Files", justify="right")
    tbl.add_column("Names", style="muted")
    for i, (prefix, files) in enumerate(groups.items(), 1):
        names = ", ".join(f.name for f in files)
        tbl.add_row(str(i), prefix, str(len(files)), names)
    console.print()
    console.print(tbl)
    console.print("  [muted]Roster matching runs during /grade.[/]")


def cmd_crit(state: dict, *args: str):
    sol = state["solutions_file"]
    if not sol:
        console.print(f"  [warning]No *solutions*.pdf found in[/] {state['cwd']}")
        return
    console.print()
    console.print(f"  [muted]solutions/criteria file:[/]")
    console.print(f"  [success]{sol.resolve()}[/]")
    try:
        import fitz as _fitz
        doc = _fitz.open(sol)
        console.print(f"  [muted]{len(doc)} page{'s' if len(doc) != 1 else ''}[/]")
        doc.close()
    except Exception:
        pass


def cmd_model(state: dict, *args: str):
    config = state["config"]
    current = state["model"]
    console.print()
    console.print(f"  [muted]Current:[/] [bold]{_model_label(current)}[/]")
    console.print()
    tbl = Table(show_header=False, box=None, padding=(0, 2), pad_edge=False)
    for i, (name, desc) in enumerate(KNOWN_MODELS, 1):
        marker = "[info]*[/]" if name == current else " "
        tbl.add_row(f"  {marker}", f"[muted]{i}.[/]", desc)
    tbl.add_row("   ", f"[muted]{len(KNOWN_MODELS)+1}.[/]", "custom — type your own")
    console.print(tbl)
    choice = Prompt.ask("  [bold]Choose[/]", default=str(1 if current is None else
                        next((i+1 for i, (m, _) in enumerate(KNOWN_MODELS) if m == current), len(KNOWN_MODELS)+1)))
    choice = (choice or "").strip()

    selected: str | None
    if choice.isdigit() and 1 <= int(choice) <= len(KNOWN_MODELS):
        selected = KNOWN_MODELS[int(choice) - 1][0]
    elif choice.isdigit() and int(choice) == len(KNOWN_MODELS) + 1:
        custom = Prompt.ask("  [bold]Model name[/]").strip()
        if not custom:
            console.print("  [muted]Cancelled.[/]")
            return
        selected = custom
    elif choice == "":
        return
    else:
        selected = choice  # treat as custom model name

    state["model"] = selected
    console.print(f"  [success]Model →[/] [bold]{_model_label(selected)}[/]")
    if selected != config.get("default_model"):
        if Confirm.ask(f"  [muted]Save as default?[/]", default=False):
            config["default_model"] = selected
            save_config(config)
            console.print(f"  [success]Saved to[/] {CONFIG_FILE}")


def _rescan(state: dict, mode: str = "auto"):
    cwd = state["cwd"]
    state["solutions_file"] = find_solutions_file(cwd)
    if state["solutions_file"]:
        groups, resolved = find_submissions(cwd, state["solutions_file"], mode=mode)
        state["groups"] = groups
        state["grouping"] = resolved
    else:
        state["groups"] = {}
        state["grouping"] = "prefix"
    state["spreadsheet"] = find_spreadsheet(cwd)


def cmd_rescan(state: dict, *args: str):
    _rescan(state, mode="auto")
    console.print("  [success]Rescanned.[/]")
    render_header(state)


def cmd_results(state: dict, *args: str):
    """Re-render the latest results table. `/results` → compact view,
    `/results full` → per-sub-part breakdown."""
    full = any(a.lower() in ("full", "detail", "detailed") for a in args)

    grades = state.get("last_grades")
    rubric = state.get("last_rubric")
    name_map = state.get("last_name_map")

    if not grades or not rubric:
        loaded = load_results(state["cwd"])
        if not loaded:
            console.print(
                f"  [warning]No saved results in {state['cwd']}.[/] "
                "Run /grade first, or cd into a directory with "
                f"{RESULTS_FILENAME}."
            )
            return
        grades = loaded["grades"]
        rubric = loaded["rubric"]
        name_map_raw = loaded.get("name_map")
        name_map = None
        if name_map_raw:
            name_map = {k: (tuple(v) if v else None)
                        for k, v in name_map_raw.items()}
        state["last_grades"] = grades
        state["last_rubric"] = rubric
        state["last_detected_names"] = loaded.get("detected_names", {})
        state["last_name_map"] = name_map
        console.print(f"  [muted]Loaded results from[/] {state['cwd'] / RESULTS_FILENAME}")

    has_subs = any("." in qid for qid in rubric)
    render_results(grades, rubric, name_map, compact=(not full and has_subs))
    if has_subs and not full:
        console.print("  [muted]Use [bold]/results full[/muted] for per-sub-part breakdown.[/]")
    elif has_subs and full:
        console.print("  [muted]Use [bold]/results[/muted] for a compact view.[/]")


def cmd_group(state: dict, *args: str):
    current = state.get("grouping", "prefix")
    console.print()
    console.print(f"  [muted]Current:[/] [bold]{current}[/]")
    console.print(
        "  [muted]Modes:[/] "
        "[bold]prefix[/] (group by name, strip trailing digits) · "
        "[bold]per-file[/] (each file = one student) · "
        "[bold]auto[/] (detect from filenames)"
    )
    choice = Prompt.ask("  [bold]Mode[/]", choices=["prefix", "per-file", "auto"],
                        default=current)
    _rescan(state, mode=choice)
    console.print(f"  [success]Grouping →[/] [bold]{state['grouping']}[/]  "
                  f"([bold]{len(state['groups'])}[/] students)")
    render_header(state)


def run_grading(state: dict, *args: str):
    sol = state["solutions_file"]
    groups = state["groups"]
    if not sol:
        console.print("  [error]Cannot grade: no solutions file.[/] Run /crit to see expectations.")
        return
    if not groups:
        console.print("  [error]Cannot grade: no student submissions.[/]")
        return

    model = state["model"]
    timeout = int(state["config"].get("codex_timeout_seconds", 1800))
    console.print(
        f"\n  Grading [bold]{len(groups)}[/] students with model "
        f"[bold]{_model_label(model)}[/]"
    )
    if not Confirm.ask("  [bold]Proceed?[/]", default=True):
        console.print("  [muted]Aborted.[/]")
        return

    with tempfile.TemporaryDirectory(prefix="autochecker_") as tmp:
        tmpdir = Path(tmp)
        with console.status("[info]Rendering solutions...[/]", spinner="dots"):
            solutions_paths = render_pdf_pages(sol, tmpdir, "solutions")

        # ── Detect rubric from the solutions file ────────────────────────
        with console.status("[info]Reading rubric from solutions...[/]", spinner="dots"):
            rubric = detect_rubric(model, timeout, solutions_paths, tmpdir)
        if rubric is None:
            console.print(
                "  [warning]Could not auto-detect rubric from solutions.[/] "
                f"Falling back to default {len(DEFAULT_RUBRIC)}-question rubric."
            )
            rubric = dict(DEFAULT_RUBRIC)

        # Group by top-level question for compact display when sub-parts exist.
        by_top: dict[str, list[tuple[str, int]]] = defaultdict(list)
        for qid, pts in rubric.items():
            top = qid.split(".", 1)[0]
            by_top[top].append((qid, pts))
        has_subs = any(len(parts) > 1 for parts in by_top.values())

        if has_subs:
            rubric_preview = "  ".join(
                f"Q{top}: {sum(p for _, p in parts)}pts"
                f"{' (' + str(len(parts)) + ' parts)' if len(parts) > 1 else ''}"
                for top, parts in by_top.items()
            )
            header = (
                f"  [success]Rubric:[/] [bold]{len(by_top)}[/] questions · "
                f"[bold]{len(rubric)}[/] atomic items · "
                f"[bold]{sum(rubric.values())}[/] points total"
            )
        else:
            rubric_preview = "  ".join(f"{qid}:{pts}" for qid, pts in rubric.items())
            header = (
                f"  [success]Rubric:[/] [bold]{len(rubric)}[/] questions · "
                f"[bold]{sum(rubric.values())}[/] points total"
            )
        console.print(header)
        console.print(f"          [muted]{rubric_preview}[/]")
        if not Confirm.ask("  [bold]Use this rubric?[/]", default=True):
            override = prompt_rubric_override()
            if override is None:
                console.print("  [muted]Aborted.[/]")
                return
            rubric = override
            by_top = defaultdict(list)
            for qid, pts in rubric.items():
                by_top[qid.split(".", 1)[0]].append((qid, pts))
            console.print(
                f"  [success]Manual rubric:[/] [bold]{len(by_top)}[/] questions · "
                f"[bold]{len(rubric)}[/] atomic items · "
                f"[bold]{sum(rubric.values())}[/] points total"
            )

        grades = {}
        detected_names: dict[str, str | None] = {}
        console.print()
        progress = Progress(
            SpinnerColumn("dots"),
            AnimatedPhaseColumn(),
            BarColumn(bar_width=30, complete_style="cyan", finished_style="green"),
            TaskProgressColumn(),
            TextColumn("[muted]|[/]"),
            TimeElapsedColumn(),
            console=console,
            transient=False,
        )
        thought_line = Text("", style="muted", no_wrap=True, overflow="ellipsis")
        with Live(Group(progress, thought_line), console=console,
                  refresh_per_second=8, transient=False):
            task = progress.add_task("Grading", total=len(groups))
            for prefix, files in groups.items():
                thought_line.plain = ""
                counters: dict = {}

                def on_event(event, _prefix=prefix, _counters=counters):
                    # Called both during name detection and grading. We read
                    # `_label[0]` so the phase prefix can update between stages.
                    phase = event_phase(event, _counters)
                    if phase:
                        progress.update(
                            task,
                            description=f"{_label[0]} [bold]{_prefix}[/] · [muted]{phase}[/]",
                        )
                    thought = event_thought(event)
                    if thought:
                        thought_line.plain = f"    ↳ {thought}"

                _label = ["Reading"]  # mutable so closure picks up new prefix
                progress.update(task,
                                description=f"Reading [bold]{prefix}[/] · starting")

                try:
                    submission_paths = materialize_submission(files, tmpdir, prefix)
                except Exception as e:
                    err_msg = str(e)[:160]
                    progress.console.print(f"  [error]Failed {prefix}:[/] {err_msg}")
                    grades[prefix] = {"scores": {qid: 0 for qid in rubric},
                                      "total": 0, "notes": f"ERROR: {err_msg}"}
                    detected_names[prefix] = None
                    progress.advance(task)
                    continue

                # ── Pre-check: read a signature off the pages ───────────────
                name = detect_student_name(
                    model, timeout, submission_paths, tmpdir, prefix,
                    on_event=on_event,
                )
                detected_names[prefix] = name
                if name:
                    name_tag = f" · [bold cyan]{name}[/]"
                    progress.console.print(f"  [muted]signed:[/] [bold cyan]{name}[/]  "
                                           f"[muted](from {prefix})[/]")
                else:
                    name_tag = " · [muted]unsigned[/]"

                # ── Grading pass ────────────────────────────────────────────
                _label[0] = "Grading"
                counters.clear()
                thought_line.plain = ""
                progress.update(
                    task,
                    description=f"Grading [bold]{prefix}[/]{name_tag} · starting",
                )
                # Update on_event closure to include the name_tag going forward.
                def on_event_grade(event, _prefix=prefix, _counters=counters,
                                   _name_tag=name_tag):
                    phase = event_phase(event, _counters)
                    if phase:
                        progress.update(
                            task,
                            description=(
                                f"Grading [bold]{_prefix}[/]{_name_tag} · "
                                f"[muted]{phase}[/]"
                            ),
                        )
                    thought = event_thought(event)
                    if thought:
                        thought_line.plain = f"    ↳ {thought}"

                try:
                    result = grade_student(model, timeout, solutions_paths, prefix,
                                           submission_paths, tmpdir, rubric,
                                           on_event=on_event_grade)
                except Exception as e:
                    err_msg = str(e)[:160]
                    progress.console.print(f"  [error]Failed {prefix}:[/] {err_msg}")
                    result = {"scores": {qid: 0 for qid in rubric}, "total": 0,
                              "notes": f"ERROR: {err_msg}"}
                result["detected_name"] = name
                grades[prefix] = result
                progress.advance(task)
            thought_line.plain = ""
            progress.update(task, description="[success]Grading complete[/]")

        spreadsheet = state["spreadsheet"]
        name_map = None
        roster_sheet = None
        roster_cols = None

        if spreadsheet:
            console.print(f"\n  [info]Spreadsheet:[/] {spreadsheet.name}")
            if Confirm.ask("  [bold]Match names & fill scores?[/]", default=True):
                with console.status("[info]Matching names to roster...[/]", spinner="dots"):
                    roster, roster_sheet, roster_cols = read_roster(spreadsheet)
                    name_map = match_names(
                        model, timeout, list(groups.keys()), roster, tmpdir,
                        signed_names=detected_names,
                    )
                console.print(
                    f"  [muted]Sheet:[/] [bold]{roster_sheet}[/]  "
                    f"[muted]cols:[/] name={_col_letter(roster_cols[0])} "
                    f"surname={_col_letter(roster_cols[1])} "
                    f"score={_col_letter(roster_cols[2])}  "
                    f"[muted]({len(roster)} rows)[/]"
                )

        # Cache everything in state for /results re-renders without re-grading.
        state["last_rubric"] = rubric
        state["last_grades"] = grades
        state["last_detected_names"] = detected_names
        state["last_name_map"] = name_map

        # Persist to disk so the user can re-open later or copy elsewhere.
        try:
            json_path, csv_path = save_results(
                state["cwd"], rubric, grades, detected_names, name_map,
            )
            console.print(
                f"\n  [success]Saved:[/] [bold]{json_path.name}[/], "
                f"[bold]{csv_path.name}[/]  [muted](in {state['cwd']})[/]"
            )
        except OSError as e:
            console.print(f"  [warning]Could not save results:[/] {e}")

        # Default to compact view if there are sub-parts; otherwise flat.
        has_subs = any("." in qid for qid in rubric)
        scores_by_name = render_results(grades, rubric, name_map,
                                        compact=has_subs)
        if has_subs:
            console.print(
                "  [muted]Compact view shown. "
                "Use [bold]/results full[/] for per-sub-part breakdown.[/]"
            )

        if name_map:
            unmatched = [p for p in grades if name_map.get(p) is None]
            if unmatched:
                console.print(
                    Panel(
                        f"[warning]{len(unmatched)} unmatched:[/] {', '.join(unmatched)}",
                        border_style="yellow",
                    )
                )

        if name_map and scores_by_name and spreadsheet and roster_sheet and roster_cols:
            if Confirm.ask(f"\n  [bold]Write {len(scores_by_name)} scores to[/] {spreadsheet.name}?", default=True):
                filled = write_scores(spreadsheet, scores_by_name, roster_sheet, roster_cols)
                console.print(f"  [success]Done![/] Wrote {filled} scores.\n")
            else:
                console.print("  [muted]Skipped.[/]\n")
        else:
            console.print()


def _top_level_id(qid: str) -> str:
    return qid.split(".", 1)[0]


def _aggregate_by_top_level(rubric: dict[str, int]) -> dict[str, int]:
    """Collapse a rubric with sub-parts into one entry per top-level question."""
    out: dict[str, int] = {}
    for qid, pts in rubric.items():
        top = _top_level_id(qid)
        out[top] = out.get(top, 0) + pts
    return out


def render_results(grades: dict, rubric: dict[str, int],
                   name_map: dict | None = None,
                   compact: bool = False) -> dict[tuple[str, str], int]:
    """Render the score table.

    `compact=True` aggregates sub-parts (e.g. 1.1 + 1.2 + …) into one column
    per top-level question. Useful when the rubric has many sub-parts and
    the full breakdown would overflow the terminal.
    """
    has_matches = name_map is not None
    has_signed = any(r.get("detected_name") for r in grades.values())
    total_max = sum(rubric.values())

    if compact:
        display_rubric = _aggregate_by_top_level(rubric)
    else:
        display_rubric = rubric

    title = "Results" + (" (compact)" if compact else "")
    table = Table(border_style="cyan", title=title, title_style="bold cyan", show_lines=True)
    table.add_column("#", style="muted", width=3, justify="right")
    table.add_column("Student", style="bold", no_wrap=True)
    if has_signed:
        table.add_column("Signed as", no_wrap=True)
    if has_matches:
        table.add_column("Matched To", no_wrap=True)

    for qid in display_rubric:
        col_label = f"Q{qid}" if qid and qid[0].isdigit() else qid
        table.add_column(col_label, justify="center",
                         width=max(4, len(col_label) + 2))
    table.add_column("Total", justify="center", style="bold", width=7)

    scores_by_name = {}
    for i, (prefix, result) in enumerate(grades.items(), 1):
        total = result.get("total", 0)
        scores = result.get("scores", {})
        detected = result.get("detected_name")

        # Build the per-displayed-column score cell.
        q_cells = []
        if compact:
            # Aggregate sub-part scores into top-level buckets.
            top_scores: dict[str, int] = defaultdict(int)
            for qid in rubric:
                top_scores[_top_level_id(qid)] += int(scores.get(qid, 0))
            for top, mx in display_rubric.items():
                s = top_scores.get(top, 0)
                q_cells.append(f"[{score_style(s, mx)}]{s}[/]/{mx}")
        else:
            for qid, mx in rubric.items():
                s = int(scores.get(qid, 0))
                q_cells.append(f"[{score_style(s, mx)}]{s}[/]/{mx}")

        style = score_style(total, total_max)
        total_cell = f"[{style}]{total}[/]/{total_max}"

        row = [str(i), prefix]
        if has_signed:
            row.append(f"[cyan]{detected}[/]" if detected else "[muted]—[/]")
        if has_matches:
            match = name_map.get(prefix)
            if match:
                row.append(f"{match[0]} {match[1]}")
                scores_by_name[match] = total
            else:
                row.append("[warning]no match[/]")
        row.extend(q_cells)
        row.append(total_cell)
        table.add_row(*row)

    console.print()
    console.print(table)
    return scores_by_name


RESULTS_FILENAME = "autochecker_results.json"
RESULTS_CSV_FILENAME = "autochecker_results.csv"


def save_results(cwd: Path, rubric: dict[str, int], grades: dict,
                 detected_names: dict[str, str | None],
                 name_map: dict | None) -> tuple[Path, Path]:
    """Persist full grading results to JSON + CSV in the grading directory.
    Returns (json_path, csv_path)."""
    json_path = cwd / RESULTS_FILENAME
    csv_path = cwd / RESULTS_CSV_FILENAME

    serialisable_name_map = None
    if name_map:
        serialisable_name_map = {
            k: (list(v) if v else None) for k, v in name_map.items()
        }

    payload = {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "version": VERSION,
        "rubric": rubric,
        "detected_names": detected_names,
        "name_map": serialisable_name_map,
        "grades": grades,
    }
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2))

    # CSV: one row per student, columns = prefix, signed_name, match_name,
    # <qid>, <qid>, ..., total
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        header = ["prefix", "signed_as", "matched_to"]
        header.extend(rubric.keys())
        header.append("total")
        header.append("notes")
        writer.writerow(header)
        for prefix, result in grades.items():
            signed = detected_names.get(prefix) or ""
            matched = ""
            if name_map and name_map.get(prefix):
                matched = " ".join(name_map[prefix])
            scores = result.get("scores") or {}
            row = [prefix, signed, matched]
            row.extend(scores.get(qid, 0) for qid in rubric)
            row.append(result.get("total", 0))
            row.append(result.get("notes", ""))
            writer.writerow(row)
    return json_path, csv_path


def load_results(cwd: Path) -> dict | None:
    path = cwd / RESULTS_FILENAME
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    return data


# ── Main ─────────────────────────────────────────────────────────────

COMMANDS = {
    "/help": cmd_help,
    "/?": cmd_help,
    "/students": cmd_students,
    "/crit": cmd_crit,
    "/model": cmd_model,
    "/group": cmd_group,
    "/grade": run_grading,
    "/results": cmd_results,
    "/rescan": cmd_rescan,
    "/refresh": cmd_rescan,
    "/status": lambda s, *a: render_header(s),
}


def build_state() -> dict:
    config = load_config()
    cwd = Path.cwd()
    solutions_file = find_solutions_file(cwd)
    if solutions_file:
        groups, grouping = find_submissions(cwd, solutions_file, mode="auto")
    else:
        groups, grouping = {}, "prefix"
    spreadsheet = find_spreadsheet(cwd)
    return {
        "config": config,
        "cwd": cwd,
        "model": config.get("default_model"),
        "solutions_file": solutions_file,
        "groups": groups,
        "grouping": grouping,
        "spreadsheet": spreadsheet,
    }


def main():
    check_codex_available()
    state = build_state()
    render_header(state)

    while True:
        try:
            raw = Prompt.ask("\n[bold cyan]›[/]")
        except (EOFError, KeyboardInterrupt):
            console.print()
            return

        cmd = (raw or "").strip()
        if not cmd:
            continue
        if cmd in ("/quit", "/exit", "/q", "quit", "exit"):
            return
        parts = cmd.split()
        head = parts[0] if parts[0].startswith("/") else "/" + parts[0]
        args = parts[1:]
        handler = COMMANDS.get(head)
        if handler is None:
            console.print(f"  [warning]Unknown command:[/] {cmd}   "
                          f"[muted](try /help)[/]")
            continue
        try:
            handler(state, *args)
        except KeyboardInterrupt:
            console.print("\n  [muted]Interrupted.[/]")


if __name__ == "__main__":
    main()
